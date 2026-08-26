import csv
import io
from datetime import datetime, timezone
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.receipt import Receipt
from app.models.user import User
from app.models.enums import ReceiptStatus, ReceiptCategory


class ExportService:
    """
    Service layer for exporting receipts in CSV and Excel formats.
    Generates downloads in-memory without creating temporary files on disk.
    """

    @staticmethod
    def _query_receipts(
        db: Session,
        status: Optional[str] = None,
        user_id: Optional[str] = None,
        category: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
    ) -> List[Receipt]:
        """Fetch receipts with optional status, user, category, and date filters applied."""
        query = (
            db.query(Receipt)
            .options(joinedload(Receipt.user), joinedload(Receipt.reviewed_by_user))
        )

        if status and status.strip() and status.strip().upper() != "ALL":
            try:
                st_enum = ReceiptStatus(status.strip().upper())
                query = query.filter(Receipt.status == st_enum)
            except ValueError:
                pass

        if user_id and user_id.strip() and user_id.strip().upper() != "ALL":
            query = query.filter(Receipt.user_id == user_id.strip())

        if category and category.strip() and category.strip().upper() != "ALL":
            try:
                cat_enum = ReceiptCategory(category.strip())
                query = query.filter(Receipt.category == cat_enum)
            except ValueError:
                pass

        if from_date and from_date.strip():
            query = query.filter(Receipt.receipt_date >= from_date.strip())
        if to_date and to_date.strip():
            query = query.filter(Receipt.receipt_date <= to_date.strip())

        return query.order_by(Receipt.submitted_at.desc()).all()

    @staticmethod
    def generate_csv(
        db: Session,
        status: Optional[str] = None,
        user_id: Optional[str] = None,
        category: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
    ) -> io.StringIO:
        """
        Generates an in-memory CSV stream of receipts.
        """
        receipts = ExportService._query_receipts(db, status, user_id, category, from_date, to_date)
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        # Header row
        writer.writerow([
            "Receipt ID",
            "Submitted By",
            "User Email",
            "Title",
            "Amount (AUD)",
            "Receipt Date",
            "Category",
            "Notes",
            "File Name",
            "Status",
            "Submitted At",
            "Reviewed At",
            "Reviewed By",
            "Review Comment",
        ])

        for r in receipts:
            sub_time_str = r.submitted_at.strftime("%Y-%m-%d %I:%M %p") if r.submitted_at else ""
            rev_time_str = r.reviewed_at.strftime("%Y-%m-%d %I:%M %p") if r.reviewed_at else ""
            writer.writerow([
                r.id,
                r.user.full_name if r.user else "",
                r.user.email if r.user else "",
                r.title,
                f"{float(r.amount):.2f}",
                r.receipt_date,
                r.category.value if r.category else "",
                r.notes or "",
                r.original_file_name,
                r.status.value if r.status else "",
                sub_time_str,
                rev_time_str,
                r.reviewed_by_user.full_name if r.reviewed_by_user else "",
                r.review_comment or "",
            ])

        output.seek(0)
        return output

    @staticmethod
    def generate_excel(
        db: Session,
        status: Optional[str] = None,
        user_id: Optional[str] = None,
        category: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
    ) -> io.BytesIO:
        """
        Generates an in-memory Excel (.xlsx) workbook of receipts with formatting.
        """
        receipts = ExportService._query_receipts(db, status, user_id, category, from_date, to_date)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Receipts Export"

        headers = [
            "Receipt ID",
            "Submitted By",
            "User Email",
            "Title",
            "Amount (AUD)",
            "Receipt Date",
            "Category",
            "Notes",
            "File Name",
            "Status",
            "Submitted At",
            "Reviewed At",
            "Reviewed By",
            "Review Comment",
        ]
        ws.append(headers)

        # Style header row (Navy blue background with bold white text)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1'),
        )

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Add data rows
        row_font = Font(name="Calibri", size=10)
        for r in receipts:
            sub_time_str = r.submitted_at.strftime("%Y-%m-%d %I:%M %p") if r.submitted_at else ""
            rev_time_str = r.reviewed_at.strftime("%Y-%m-%d %I:%M %p") if r.reviewed_at else ""
            ws.append([
                r.id,
                r.user.full_name if r.user else "",
                r.user.email if r.user else "",
                r.title,
                float(r.amount),
                r.receipt_date,
                r.category.value if r.category else "",
                r.notes or "",
                r.original_file_name,
                r.status.value if r.status else "",
                sub_time_str,
                rev_time_str,
                r.reviewed_by_user.full_name if r.reviewed_by_user else "",
                r.review_comment or "",
            ])

        # Format data cells and adjust column widths
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = row_font
                cell.border = thin_border
                # Format amount column as currency/decimal
                if cell.column == 5:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column width with a minimum width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


export_service = ExportService()
