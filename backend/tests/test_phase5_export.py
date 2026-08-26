import io
import uuid
import openpyxl
from app.models.user import User
from app.models.receipt import Receipt
from app.models.enums import UserRole, ReceiptStatus, ReceiptCategory
from app.core.security import hash_password


def test_exports(client, db_session):
    suffix = uuid.uuid4().hex[:6]
    admin_email = f"admin.export.{suffix}@example.com"
    user_email = f"user.export.{suffix}@example.com"

    # Create Admin
    admin = User(
        full_name="Admin Export",
        email=admin_email,
        hashed_password=hash_password("AdminPassword123"),
        role=UserRole.ADMIN,
    )
    # Create User
    user = User(
        full_name="User Export",
        email=user_email,
        hashed_password=hash_password("Password123"),
        role=UserRole.USER,
    )
    db_session.add_all([admin, user])
    db_session.commit()

    # Create an APPROVED receipt in DB
    approved_receipt = Receipt(
        user_id=user.id,
        title="Client Dinner Export Test",
        amount=4500.00,
        receipt_date="2026-08-24",
        category=ReceiptCategory.CLIENT_MEETINGS,
        notes="Export test notes",
        original_file_name="receipt.pdf",
        stored_file_name="stored_receipt.pdf",
        file_path="uploads/test.pdf",
        content_type="application/pdf",
        status=ReceiptStatus.APPROVED,
        review_comment="Approved for export",
        reviewed_by_id=admin.id,
    )
    db_session.add(approved_receipt)
    db_session.commit()

    # Login User
    res_user = client.post('/api/auth/login', json={'email': user_email, 'password': 'Password123'})
    user_token = res_user.json()['access_token']

    # Login Admin
    res_admin = client.post('/api/auth/login', json={'email': admin_email, 'password': 'AdminPassword123'})
    admin_token = res_admin.json()['access_token']

    # 1. Non-admin export CSV -> 403 Forbidden
    user_csv = client.get('/api/admin/receipts/export/csv', headers={'Authorization': f'Bearer {user_token}'})
    assert user_csv.status_code == 403

    # 2. Admin export CSV -> 200 OK
    admin_csv = client.get('/api/admin/receipts/export/csv', headers={'Authorization': f'Bearer {admin_token}'})
    assert admin_csv.status_code == 200
    assert 'text/csv' in admin_csv.headers['content-type']
    assert 'attachment;' in admin_csv.headers['content-disposition']
    assert 'Receipt ID,Submitted By' in admin_csv.text

    # 3. Admin export Excel -> 200 OK
    admin_excel = client.get('/api/admin/receipts/export/excel', headers={'Authorization': f'Bearer {admin_token}'})
    assert admin_excel.status_code == 200
    assert 'spreadsheetml' in admin_excel.headers['content-type']
    assert 'attachment;' in admin_excel.headers['content-disposition']

    # Verify excel workbook is valid
    wb = openpyxl.load_workbook(io.BytesIO(admin_excel.content))
    assert 'Receipts Export' in wb.sheetnames
    ws = wb.active
    assert ws.max_row >= 2
